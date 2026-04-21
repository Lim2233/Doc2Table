import os
import glob
import json
import logging
import argparse
from openpyxl import load_workbook

# 配置日志格式
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def fill_table_from_json(worksheet, column_names, json_data, start_row=2):
    """
    根据 JSON 数据填充工作表

    :param worksheet:    openpyxl 工作表对象
    :param column_names: 第一行的列名列表
    :param json_data:    JSON 数据列表，每个元素为字典
    :param start_row:    数据填充的起始行号，默认为 2
    """
    col_key_map = {
        idx: name for idx, name in enumerate(column_names, start=1) if name is not None
    }

    for row_offset, record in enumerate(json_data):
        row_idx = start_row + row_offset
        for col_idx, key in col_key_map.items():
            if key in record:
                worksheet.cell(row=row_idx, column=col_idx, value=record[key])


def merge_json_data(json_files):
    """
    读取多个 JSON 文件并将数据合并为一个列表

    :param json_files: JSON 文件路径列表
    :return: 合并后的数据列表
    """
    merged = []
    for json_path in json_files:
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            if isinstance(data, list):
                merged.extend(data)
                logger.info(f"加载 {os.path.basename(json_path)}：{len(data)} 条记录")
            else:
                logger.warning(f"文件 {json_path} 内容不是列表，跳过")
        except Exception as e:
            logger.error(f"读取 JSON 文件 {json_path} 失败：{e}")
    return merged


def find_single_file(folder, extension, description):
    """
    在指定文件夹中查找指定扩展名的文件，要求有且仅有一个。

    :param folder:      文件夹路径
    :param extension:   文件扩展名（如 ".json", ".xlsx"）
    :param description: 用于错误提示的描述文字
    :return:            找到的唯一文件的完整路径
    """
    pattern = os.path.join(folder, f"*{extension}")
    files = glob.glob(pattern)

    if not files:
        raise FileNotFoundError(f"在文件夹 '{folder}' 中未找到 {description} 文件（{extension}）。")
    if len(files) > 1:
        raise ValueError(f"文件夹 '{folder}' 中包含多个 {description} 文件，但只允许存在一个。")
    return files[0]


def main():
    """命令行入口"""
    parser = argparse.ArgumentParser(
        description="根据多个 JSON 数据文件合并后填充单个 Excel 模板",
        usage="python %(prog)s <数据文件夹> <模板文件夹> <输出文件夹> [选项]"
    )
    parser.add_argument("data_folder", help="存放 JSON 数据文件的文件夹（可包含多个 .json 文件）")
    parser.add_argument("template_folder", help="存放 Excel 模板的文件夹（仅一个 .xlsx 文件）")
    parser.add_argument("output_folder", help="输出文件夹路径")
    parser.add_argument("--sheet", default=None, help="工作表名称，默认使用活动工作表")
    parser.add_argument("--header-row", type=int, default=1, help="列名所在行号，默认 1")
    parser.add_argument("--start-row", type=int, default=2, help="数据填充起始行号，默认 2")
    parser.add_argument("--output-name", default=None, help="输出文件名（不含路径），默认基于模板名生成")

    args = parser.parse_args()

    # 检查模板文件夹中是否恰好有一个 Excel 模板文件
    try:
        template_path = find_single_file(args.template_folder, ".xlsx", "Excel 模板")
    except Exception as e:
        logger.error(str(e))
        return

    # 获取数据文件夹中所有 JSON 文件
    json_pattern = os.path.join(args.data_folder, "*.json")
    json_files = glob.glob(json_pattern)

    if not json_files:
        logger.error(f"在文件夹 '{args.data_folder}' 中未找到任何 JSON 文件。")
        return

    logger.info(f"找到 {len(json_files)} 个 JSON 文件，开始合并数据...")

    # 合并所有 JSON 数据
    merged_data = merge_json_data(json_files)
    if not merged_data:
        logger.warning("合并后的数据为空，将生成空白输出文件。")
    logger.info(f"合并完成，共 {len(merged_data)} 条记录。")

    # 确定输出文件名
    if args.output_name:
        output_filename = args.output_name
    else:
        base = os.path.splitext(os.path.basename(template_path))[0]
        output_filename = f"{base}_filled.xlsx"

    output_path = os.path.join(args.output_folder, output_filename)

    # 确保输出文件夹存在
    os.makedirs(args.output_folder, exist_ok=True)

    # 执行填充
    try:
        wb = load_workbook(template_path, data_only=True)
        ws = wb[args.sheet] if args.sheet else wb.active

        # 读取列名
        header_cells = list(ws.iter_rows(min_row=args.header_row, max_row=args.header_row, values_only=True))
        if not header_cells:
            logger.error(f"模板文件 {template_path} 第 {args.header_row} 行为空，无法处理。")
            wb.close()
            return

        column_names = header_cells[0]

        # 填充数据
        fill_table_from_json(ws, column_names, merged_data, start_row=args.start_row)

        # 保存
        wb.save(output_path)
        logger.info(f"已成功生成文件：{output_path}")

        wb.close()
    except Exception as e:
        logger.error(f"处理过程中出错：{e}")
        return

    logger.info("所有操作完成。")


if __name__ == "__main__":
    main()