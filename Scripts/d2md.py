"""
docling_batch_converter.py
批量文档转 Markdown（模块化设计，支持扩展处理器）

CLI用法为 python <脚本名> <输入文件夹> <md文件保存文件夹>
"""

import logging
import warnings
from pathlib import Path
from typing import Dict, Callable, Optional, List, Tuple, Any, Union
from dataclasses import dataclass, field

import openpyxl
from docling.document_converter import DocumentConverter

# 配置模块级日志
logger = logging.getLogger(__name__)


# ----------------------------------------------------------------------
# 默认处理器实现
# ----------------------------------------------------------------------
def txt_to_markdown(file_path: Path, encoding: str = "utf-8") -> str:
    """
    直接读取 .txt 文件，将内容包裹在 Markdown 代码块中

    Args:
        file_path: 文本文件路径
        encoding: 尝试的编码，失败时回退到 gbk

    Returns:
        Markdown 格式的字符串
    """
    try:
        content = file_path.read_text(encoding=encoding)
    except UnicodeDecodeError:
        content = file_path.read_text(encoding="gbk")
    return f"```text\n{content}\n```"


def xlsx_to_markdown(file_path: Path) -> str:
    """
    使用 openpyxl 读取 Excel，生成 Markdown 表格

    Args:
        file_path: Excel 文件路径

    Returns:
        Markdown 表格字符串
    """
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            return f"<!-- 错误：无法打开 Excel 文件 - {e} -->\n\n*文件可能已损坏或格式不支持。*"

    sheet_names = wb.sheetnames
    worksheets = wb.worksheets

    if not sheet_names and not worksheets:
        wb.close()
        return "<!-- 警告：无法读取任何工作表，文件可能严重损坏。\n请用 Excel 打开此文件，另存为新的 .xlsx 文件后重试。 -->\n\n*(无数据)*"

    md_parts = []
    # 当 sheetnames 为空但 worksheets 存在时，强制读取第一个工作表
    if not sheet_names and worksheets:
        logger.warning(f"工作表列表为空，强制读取第一个工作表: {file_path}")
        ws = worksheets[0]
        sheet_names = [ws.title or "Sheet1"]

    for sheet_name in sheet_names:
        ws = wb[sheet_name] if sheet_name else wb.worksheets[0]
        actual_name = sheet_name if sheet_name else "Sheet1"
        rows = list(ws.iter_rows(values_only=True))
        non_empty_rows = [row for row in rows if any(cell is not None for cell in row)]

        if not non_empty_rows:
            md_parts.append(f"<!-- 工作表: {actual_name} -->\n*(工作表无数据)*\n")
            continue

        max_cols = max(len(row) for row in non_empty_rows)
        aligned_rows = []
        for row in non_empty_rows:
            aligned_rows.append(list(row) + [''] * (max_cols - len(row)))

        header = aligned_rows[0]
        md_parts.append(f"<!-- 工作表: {actual_name} -->\n")
        md_parts.append("| " + " | ".join(str(cell) if cell is not None else '' for cell in header) + " |")
        md_parts.append("|" + "|".join([" --- "] * max_cols) + "|")

        for row in aligned_rows[1:]:
            md_parts.append("| " + " | ".join(str(cell) if cell is not None else '' for cell in row) + " |")

        md_parts.append("\n")

    wb.close()
    return "\n".join(md_parts)


def docling_to_markdown(file_path: Path, converter: Optional[DocumentConverter] = None) -> str:
    """
    使用 docling 将文档转换为 Markdown

    Args:
        file_path: 文档路径（docx, md 等）
        converter: 可复用的 DocumentConverter 实例

    Returns:
        Markdown 字符串
    """
    if converter is None:
        converter = DocumentConverter()
    result = converter.convert(str(file_path))
    return result.document.export_to_markdown()


# ----------------------------------------------------------------------
# 转换器类
# ----------------------------------------------------------------------
@dataclass
class ConversionStats:
    """转换统计信息"""
    total: int = 0
    success: int = 0
    failed: int = 0
    skipped: int = 0
    details: List[Dict[str, Any]] = field(default_factory=list)


class DocumentMarkdownConverter:
    """
    批量文档转 Markdown 转换器，支持自定义处理器和回调钩子
    """

    def __init__(
        self,
        input_dir: Optional[Path] = None,
        output_dir: Optional[Path] = None,
        extensions: Tuple[str, ...] = ('.docx', '.md', '.txt'),   # 固定扩展名
        recursive: bool = False,
        overwrite: bool = True,
        encoding: str = "utf-8",
        custom_handlers: Optional[Dict[str, Callable]] = None,
        logger: Optional[logging.Logger] = None,
    ):
        """
        初始化转换器

        Args:
            input_dir: 输入根目录（可为 None，之后通过 convert() 指定）
            output_dir: 输出根目录
            extensions: 需要处理的文件扩展名元组（如 .docx）
            recursive: 是否递归搜索子目录
            overwrite: 是否覆盖已有输出文件
            encoding: 文本文件读取编码
            custom_handlers: 自定义扩展名到处理函数的映射，覆盖默认处理器
            logger: 日志记录器，未提供则使用模块级 logger
        """
        self.input_dir = Path(input_dir) if input_dir else None
        self.output_dir = Path(output_dir) if output_dir else None
        self.extensions = tuple(ext.lower() for ext in extensions)
        self.recursive = recursive
        self.overwrite = overwrite
        self.encoding = encoding
        self.logger = logger or logging.getLogger(__name__)

        # 初始化默认处理器
        self._handlers: Dict[str, Callable] = {
            '.txt': self._wrap_txt_handler,
            '.xlsx': self._wrap_xlsx_handler,
        }
        # docling 处理器用于其他格式
        self._docling_converter = DocumentConverter()
        self._default_handler = self._wrap_docling_handler

        # 合并自定义处理器
        if custom_handlers:
            self._handlers.update(custom_handlers)

    # ------------------------- 内部包装方法 -------------------------
    def _wrap_txt_handler(self, file_path: Path) -> str:
        return txt_to_markdown(file_path, encoding=self.encoding)

    def _wrap_xlsx_handler(self, file_path: Path) -> str:
        return xlsx_to_markdown(file_path)

    def _wrap_docling_handler(self, file_path: Path) -> str:
        return docling_to_markdown(file_path, converter=self._docling_converter)

    # ------------------------- 公共方法 -------------------------
    def register_handler(self, extension: str, handler: Callable[[Path], str]) -> None:
        """
        注册新的文件处理器

        Args:
            extension: 文件扩展名（如 '.pdf'）
            handler: 接受 Path 参数返回 Markdown 字符串的函数
        """
        self._handlers[extension.lower()] = handler

    def convert_file(self, input_path: Path, output_path: Optional[Path] = None) -> Tuple[bool, str]:
        """
        转换单个文件

        Args:
            input_path: 源文件路径
            output_path: 输出文件路径（若不指定则根据 input_dir/output_dir 自动生成）

        Returns:
            (是否成功, 输出文件路径或错误信息)
        """
        if not input_path.is_file():
            return False, f"文件不存在: {input_path}"

        # 确定输出路径
        if output_path is None:
            if self.input_dir is None or self.output_dir is None:
                raise ValueError("未指定 output_path，且转换器未配置 input_dir/output_dir")
            rel_path = input_path.relative_to(self.input_dir)
            out_file = self.output_dir / rel_path.with_suffix('.md')
        else:
            out_file = output_path

        # 检查是否需要跳过
        if not self.overwrite and out_file.exists():
            self.logger.debug(f"跳过已存在文件: {out_file}")
            return True, "skipped"

        # 选择处理器
        ext = input_path.suffix.lower()
        handler = self._handlers.get(ext, self._default_handler)

        try:
            out_file.parent.mkdir(parents=True, exist_ok=True)
            md_content = handler(input_path)
            out_file.write_text(md_content, encoding=self.encoding)
            return True, str(out_file)
        except Exception as e:
            self.logger.error(f"转换失败 {input_path}: {e}", exc_info=True)
            return False, str(e)

    def convert(
        self,
        input_dir: Optional[Union[str, Path]] = None,
        output_dir: Optional[Union[str, Path]] = None,
        on_file_start: Optional[Callable[[Path], None]] = None,
        on_file_success: Optional[Callable[[Path, Path], None]] = None,
        on_file_failure: Optional[Callable[[Path, Exception], None]] = None,
    ) -> ConversionStats:
        """
        批量转换目录下的所有匹配文件

        Args:
            input_dir: 输入目录（覆盖实例的 input_dir）
            output_dir: 输出目录（覆盖实例的 output_dir）
            on_file_start: 文件开始转换时的回调，参数为源文件路径
            on_file_success: 文件转换成功回调，参数为源文件路径、输出文件路径
            on_file_failure: 文件转换失败回调，参数为源文件路径、异常信息

        Returns:
            ConversionStats 统计对象
        """
        if input_dir:
            input_dir = Path(input_dir)
        else:
            input_dir = self.input_dir
        if output_dir:
            output_dir = Path(output_dir)
        else:
            output_dir = self.output_dir

        if input_dir is None or output_dir is None:
            raise ValueError("必须指定 input_dir 和 output_dir")

        input_dir = Path(input_dir).resolve()
        output_dir = Path(output_dir).resolve()

        # 收集所有待处理文件
        if self.recursive:
            files = [p for p in input_dir.rglob("*") if p.is_file() and p.suffix.lower() in self.extensions]
        else:
            files = [p for p in input_dir.iterdir() if p.is_file() and p.suffix.lower() in self.extensions]

        stats = ConversionStats(total=len(files))

        if not files:
            self.logger.info(f"未找到匹配的文件: {input_dir}")
            return stats

        self.logger.info(f"找到 {len(files)} 个文件，开始转换...")

        for src in files:
            if on_file_start:
                on_file_start(src)

            # 计算相对路径并确定输出路径
            rel_path = src.relative_to(input_dir)
            out_file = output_dir / rel_path.with_suffix('.md')

            success, msg = self.convert_file(src, out_file)
            if success:
                stats.success += 1
                if msg == "skipped":
                    stats.skipped += 1
                if on_file_success:
                    on_file_success(src, out_file)
                self.logger.debug(f"转换成功: {src} -> {out_file}")
            else:
                stats.failed += 1
                if on_file_failure:
                    on_file_failure(src, msg)
                self.logger.error(f"转换失败: {src} - {msg}")

            stats.details.append({
                "source": str(src),
                "success": success,
                "message": msg,
            })

        self.logger.info(f"转换完成: 成功 {stats.success}, 失败 {stats.failed}, 跳过 {stats.skipped}")
        return stats


# ----------------------------------------------------------------------
# 便捷函数（保持向后兼容）
# ----------------------------------------------------------------------
def batch_convert_to_markdown(
    input_dir: str,
    output_dir: str,
    extensions: Tuple[str, ...] = ('.docx', '.md', '.txt')   # 固定扩展名
) -> Dict[str, int]:
    """
    原始函数风格的批量转换（向后兼容）

    Returns:
        统计信息字典: {'success': int, 'failed': int, 'total': int}
    """
    converter = DocumentMarkdownConverter(
        input_dir=Path(input_dir),
        output_dir=Path(output_dir),
        extensions=extensions,
        overwrite=True,
        recursive=False,
    )
    stats = converter.convert()
    return {"success": stats.success, "failed": stats.failed, "total": stats.total}


# ----------------------------------------------------------------------
# 命令行入口
# ----------------------------------------------------------------------
if __name__ == '__main__':
    import argparse

    parser = argparse.ArgumentParser(
        description="批量文档转 Markdown（固定处理 .docx, .md, .txt 文件）"
    )
    parser.add_argument("input", help="输入文件夹路径")
    parser.add_argument("output", help="md 文件保存文件夹路径")
    parser.add_argument("--recursive", "-r", action="store_true", help="递归搜索子目录")
    parser.add_argument("--no-overwrite", action="store_true", help="不覆盖已存在的输出文件")
    parser.add_argument("--log-level", default="INFO", choices=["DEBUG", "INFO", "WARNING", "ERROR"],
                        help="日志级别（默认 INFO）")

    args = parser.parse_args()

    logging.basicConfig(
        level=getattr(logging, args.log_level),
        format="%(asctime)s - %(levelname)s - %(message)s"
    )

    converter = DocumentMarkdownConverter(
        input_dir=Path(args.input),
        output_dir=Path(args.output),
        extensions=('.docx', '.md', '.txt'),          # 固定扩展名
        recursive=args.recursive,
        overwrite=not args.no_overwrite,
    )
    stats = converter.convert()
    print(f"统计: 成功 {stats.success}, 失败 {stats.failed}, 跳过 {stats.skipped}, 总计 {stats.total}")